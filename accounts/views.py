from django.shortcuts import render, redirect, get_object_or_404
from .forms import StudentForm, GraduateForm
from .models import Student, Graduate
from django.http import HttpResponse
from django.contrib import messages
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


# -------------------
# ホーム
# -------------------
def home(request):
    return render(request, 'home.html')


# -------------------
# 学生
# -------------------
def student_create(request):
    if request.method == 'POST':
        form = StudentForm(request.POST)
        if form.is_valid():
            obj = form.save(commit=False)

            if form.cleaned_data.get('desired_department') == "その他":
                obj.desired_department = form.cleaned_data.get('desired_department_other')

            obj.save()
            messages.success(request, "登録しました！")
            return redirect('/students/')
        else:
            print(form.errors)
    else:
        form = StudentForm()

    return render(request, 'student_form.html', {'form': form})


def student_list(request):
    students = Student.objects.all()
    return render(request, 'student_list.html', {'students': students})


def student_delete(request, id):
    student = get_object_or_404(Student, id=id)
    student.delete()
    return redirect('/students/')


def student_edit(request, id):
    student = get_object_or_404(Student, id=id)

    if request.method == 'POST':
        form = StudentForm(request.POST, instance=student)
        if form.is_valid():
            obj = form.save(commit=False)

            if form.cleaned_data.get('desired_department') == "その他":
                obj.desired_department = form.cleaned_data.get('desired_department_other')

            obj.save()
            messages.success(request, "更新しました！")
            return redirect('/students/')
        else:
            print(form.errors)
    else:
        form = StudentForm(instance=student)

    return render(request, 'student_form.html', {'form': form})


# -------------------
# 卒業生
# -------------------
def graduate_create(request):
    if request.method == 'POST':
        form = GraduateForm(request.POST)
        if form.is_valid():
            obj = form.save(commit=False)

            if form.cleaned_data.get('department') == "その他":
                obj.department = form.cleaned_data.get('department_other')

            obj.save()
            messages.success(request, "登録しました！")
            return redirect('/graduates/')
        else:
            print(form.errors)
    else:
        form = GraduateForm()

    return render(request, 'graduate_form.html', {'form': form})


def graduate_list(request):
    graduates = Graduate.objects.all()
    return render(request, 'graduate_list.html', {'graduates': graduates})


def graduate_delete(request, id):
    graduate = get_object_or_404(Graduate, id=id)
    graduate.delete()
    return redirect('/graduates/')


def graduate_edit(request, id):
    graduate = get_object_or_404(Graduate, id=id)

    if request.method == 'POST':
        form = GraduateForm(request.POST, instance=graduate)
        if form.is_valid():
            obj = form.save(commit=False)

            if form.cleaned_data.get('department') == "その他":
                obj.department = form.cleaned_data.get('department_other')

            obj.save()
            messages.success(request, "更新しました！")
            return redirect('/graduates/')
        else:
            print(form.errors)
    else:
        form = GraduateForm(instance=graduate)

    return render(request, 'graduate_form.html', {'form': form})


# -------------------
# Excel（学生）
# -------------------
def export_students_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Students'

    headers = [
        '名前', 'ふりがな', '生年月日', '学年',
        '出身', '高校', '卒業年度', '趣味', '志望診療科'
    ]
    ws.append(headers)

    # ヘッダー装飾
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # データ
    for s in Student.objects.all():
        ws.append([
            s.name,
            s.furigana,
            s.birth_date.strftime("%Y-%m-%d") if s.birth_date else "",
            s.grade,
            s.hometown,
            s.high_school,
            s.high_school_graduation_year,
            s.hobby,
            s.desired_department
        ])

    # 列幅
    widths = [15, 15, 15, 8, 15, 20, 12, 25, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    # フィルター＆固定
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=students.xlsx'
    wb.save(response)

    return response


# -------------------
# Excel（卒業生）
# -------------------
def export_graduates_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Graduates'

    headers = [
        '名前', 'ふりがな', '生年月日', '診療科',
        '高校', '大学', '卒業年', '勤務先', 'メール', '電話番号'
    ]
    ws.append(headers)

    # ヘッダー装飾
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # データ
    for g in Graduate.objects.all():
        ws.append([
            g.name,
            g.furigana,
            g.birth_date.strftime("%Y-%m-%d") if g.birth_date else "",
            g.department,
            g.high_school,
            g.university,
            g.graduation_year,
            g.hospital,
            g.email,
            g.phone
        ])

    # 列幅
    widths = [15, 15, 15, 15, 20, 20, 12, 25, 25, 15]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=graduates.xlsx'
    wb.save(response)

    return response